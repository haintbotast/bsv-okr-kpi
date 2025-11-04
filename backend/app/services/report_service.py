"""Report generation service."""

from typing import List, Optional
from sqlalchemy.orm import Session
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.models.kpi import KPI
from app.crud.kpi import kpi_crud


class ReportService:
    """Service for generating reports."""

    def generate_excel_report(
        self,
        db: Session,
        user_id: Optional[int] = None,
        year: Optional[int] = None,
        quarter: Optional[str] = None,
        status: Optional[str] = None
    ) -> BytesIO:
        """Generate Excel report for KPIs."""
        # Get KPIs
        kpis, _ = kpi_crud.get_multi(
            db,
            skip=0,
            limit=1000,
            user_id=user_id,
            year=year,
            quarter=quarter,
            status=status
        )

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "KPI Report"

        # Header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Headers
        headers = ["ID", "Title", "Year", "Quarter", "Category", "Status",
                   "Target", "Current", "Progress %", "Created", "Updated"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row, kpi in enumerate(kpis, 2):
            ws.cell(row=row, column=1, value=kpi.id)
            ws.cell(row=row, column=2, value=kpi.title)
            ws.cell(row=row, column=3, value=kpi.year)
            ws.cell(row=row, column=4, value=kpi.quarter)
            ws.cell(row=row, column=5, value=kpi.category or "N/A")
            ws.cell(row=row, column=6, value=kpi.status)
            ws.cell(row=row, column=7, value=kpi.target_value or "N/A")
            ws.cell(row=row, column=8, value=kpi.current_value or "N/A")
            ws.cell(row=row, column=9, value=kpi.progress_percentage or 0)
            ws.cell(row=row, column=10, value=kpi.created_at.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=11, value=kpi.updated_at.strftime("%Y-%m-%d"))

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def get_analytics_data(
        self,
        db: Session,
        user_id: Optional[int] = None,
        year: Optional[int] = None
    ) -> dict:
        """Get analytics data for dashboards."""
        # Get all KPIs
        kpis, total = kpi_crud.get_multi(
            db,
            skip=0,
            limit=10000,
            user_id=user_id,
            year=year
        )

        if not kpis:
            return {
                "total_kpis": 0,
                "avg_progress": 0,
                "completion_rate": 0,
                "by_status": {},
                "by_quarter": {},
                "by_category": {}
            }

        # Calculate metrics
        total_kpis = len(kpis)
        avg_progress = sum(k.progress_percentage or 0 for k in kpis) / total_kpis
        approved = sum(1 for k in kpis if k.status == "approved")
        completion_rate = (approved / total_kpis * 100) if total_kpis > 0 else 0

        # By status
        by_status = {}
        for kpi in kpis:
            by_status[kpi.status] = by_status.get(kpi.status, 0) + 1

        # By quarter
        by_quarter = {}
        for kpi in kpis:
            by_quarter[kpi.quarter] = by_quarter.get(kpi.quarter, 0) + 1

        # By category
        by_category = {}
        for kpi in kpis:
            cat = kpi.category or "Uncategorized"
            by_category[cat] = by_category.get(cat, 0) + 1

        return {
            "total_kpis": total_kpis,
            "avg_progress": round(avg_progress, 2),
            "completion_rate": round(completion_rate, 2),
            "by_status": by_status,
            "by_quarter": by_quarter,
            "by_category": by_category
        }


# Singleton instance
report_service = ReportService()
